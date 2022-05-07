#!/usr/bin/env python3
import argparse
import logging
import os
import glob
import shutil
from isort import file
import textract
import pytesseract
import warnings

from tqdm import tqdm
from multiprocessing import Pool, Lock, TimeoutError
from tempfile import TemporaryDirectory
from pyunpack import Archive, PatoolError
from email.parser import Parser as EmailParser
from email.policy import default
from openpyxl import load_workbook
from xlrd.compdoc import CompDocError
from pytesseract.pytesseract import TesseractError
from zipfile import BadZipFile

import numpy as np
import pandas as pd
from skimage import io
from skimage.transform import rotate


logger = logging.getLogger(__name__)
logger.setLevel(logging.ERROR)

consoleHandler = logging.StreamHandler()
consoleHandler.setLevel(logging.ERROR)

logger.addHandler(consoleHandler)

formatter = logging.Formatter('%(asctime)s  %(levelname)s: %(message)s')
consoleHandler.setFormatter(formatter)

zip_exts = {".7z", ".ace", ".alz", ".a", ".arc", ".arj", ".bz2", ".cab", ".Z", ".cpio", ".deb",
    ".dms", ".gz", ".lrz", ".lha", ".lzh", ".lz", ".lzma", ".lzo", ".rpm", ".rar", ".rz", ".tar",
    ".xz", ".zip", ".jar", ".zoo"}

img_exts = {".gif", ".jpg", ".jpeg", ".png", ".tiff", ".tif"}     # ".pdf"

parseable_exts = {".csv", ".doc", ".docx", ".eml", ".epub", ".gif", ".jpg", ".jpeg", ".json",
    ".html", ".htm", ".mp3", ".msg", ".odt", ".ogg", ".pdf", ".png", ".pptx", ".ps", ".rtf",
    ".tiff", ".tif", ".txt", ".wav", ".xlsx", ".xls"}

fs_lock = Lock()


def get_normalized_ext(fn):
    return os.path.splitext(fn)[1].lower()


def is_zipfile(fn):
    return get_normalized_ext(fn) in zip_exts


def is_image(fn):
    return get_normalized_ext(fn) in img_exts


def is_parseable(fn):
    return get_normalized_ext(fn) in parseable_exts


def _prefixed_fn(fn, prefix, output_dir):
    return os.path.join(output_dir, prefix + os.path.basename(fn))


def _postfixed_fn(fn, postfix):
    if postfix == 0:
        return fn
    else:
        return fn + '.' + str(postfix)


def _non_conflicting_fn(fn, output_dir):
    res_fn = os.path.join(output_dir, os.path.basename(fn))
    postfix = 0
    while os.path.isfile(_postfixed_fn(res_fn, postfix)):
        postfix += 1
    return _postfixed_fn(res_fn, postfix)


def _init_pool_processes(the_lock):
    global fs_lock
    fs_lock = the_lock


def _get_empty_file_status_dict():
    return {
        "input_path": [],
        "input_filename": [],
        "output_filename": [],
        "status": [],
        "details": []
    }


def _append_file_status(file_status, file_status_csv):
    logger.info(f"Appending {len(file_status['input_path'])} records to file status csv")
    if os.path.isfile(file_status_csv) and os.path.getsize(file_status_csv) > 0:
        # Note that this does NOT really mean that the file has a correct header but we judge it
        #   not worth it to open the file to validate it
        pd \
            .DataFrame(file_status) \
            .to_csv(file_status_csv, index=False, header=False, mode='a')
    else:
        pd \
            .DataFrame(file_status) \
            .to_csv(file_status_csv, index=False, header=True, mode='w')


def _add_file_status(file_status, input_path, input_filename, output_filename, status, details):
    file_status["input_path"].append(input_path)
    file_status["input_filename"].append(input_filename)
    file_status["output_filename"].append(output_filename)
    file_status["status"].append(status)
    file_status["details"].append(details)


def _file_op(args):
    fn, prefix, output_dir, file_op_lambda = args
    res = file_op_lambda(fn, prefix, output_dir)
    return fn, res


def process_dir(
    input_dir,
    output_dir,
    file_op_lambda,
    num_pool_procs=20,
    pool_timeout=60,
    temp_dir=None,
    prefix="",
    do_not_unzip=False,
    already_processed=set(),
    file_status_csv="./textify_file_status.csv",
    file_status_save_period=1000
):
    all_files = glob.glob(os.path.join(input_dir, "**"), recursive=True)

    zip_files = [fn for fn in all_files if is_zipfile(fn)]
    file_status = _get_empty_file_status_dict()
    if len(zip_files) > 0:
        logger.info(f"Processing archives in {input_dir} with prefix \"{prefix}\"...")
        for zip_fn in tqdm(zip_files, leave=False, desc=prefix):
            with TemporaryDirectory(dir=temp_dir) as temp_dir_name:
                try:
                    Archive(zip_fn).extractall(os.path.join(temp_dir_name))
                except PatoolError:
                    logger.warning(f"Failed to extract {zip_fn} with prefix \"{prefix}\"")
                    cur_file_status = "Unzip failed"
                    continue
                except RuntimeError:
                    logger.warning(f"Encrypted: {zip_fn} with prefix \"{prefix}\"")
                    cur_file_status = "Zip file encrypted"
                    continue
                cur_file_status = "Unzipped successfully"
                _add_file_status(
                    file_status,
                    prefix,
                    os.path.basename(zip_fn),
                    "",
                    cur_file_status,
                    "")
                zip_file_status = process_dir(
                    temp_dir_name,
                    output_dir,
                    file_op_lambda,
                    num_pool_procs=num_pool_procs,
                    pool_timeout=pool_timeout,
                    prefix=prefix + os.path.basename(zip_fn) + "/",
                    temp_dir=temp_dir,
                    do_not_unzip=do_not_unzip,
                    already_processed=already_processed,
                    file_status_csv=file_status_csv,
                    file_status_save_period=file_status_save_period)
                for key, val in file_status.items():
                    file_status[key] = val + zip_file_status[key]
                if len(file_status["input_path"]) > file_status_save_period:
                    _append_file_status(file_status, file_status_csv)
                    file_status = _get_empty_file_status_dict()
    logger.info(f"Processing files in {input_dir} with prefix \"{prefix}\"...")
    with Pool(
        processes=num_pool_procs,
        maxtasksperchild=1,
        initializer=_init_pool_processes,
        initargs=(fs_lock,)
    ) as pool:
        job_args = [
            (fn, prefix, output_dir, file_op_lambda)
            for fn in all_files
            if
                not is_zipfile(fn) and
                os.path.isfile(fn) and
                not (prefix + os.path.basename(fn)) in already_processed
        ]
        pool_res = pool.imap_unordered(_file_op, job_args)
        non_timed_out = set()
        with tqdm(total=len(job_args), leave=False) as pbar:
            while True:
                try:
                    cur_res = pool_res.next(pool_timeout)
                except StopIteration:
                    break
                except TimeoutError:
                    logger.warning(f"Timeout during processing prefix \"{prefix}\"")
                    break
                input_fn, (output_fn, res, details) = cur_res
                non_zip_fn = os.path.basename(input_fn)
                non_timed_out.add(non_zip_fn)
                if res is not None:
                    logger.warning(f"{res} during processing {non_zip_fn} with prefix \"{prefix}\"")
                    cur_file_status = res
                    cur_file_details = details
                else:
                    cur_file_status = "Success"
                    cur_file_details = ""
                _add_file_status(
                    file_status,
                    prefix,
                    non_zip_fn,
                    output_fn,
                    cur_file_status,
                    cur_file_details)
                if len(file_status["input_path"]) > file_status_save_period:
                    _append_file_status(file_status, file_status_csv)
                    file_status = _get_empty_file_status_dict()
                pbar.update(1)
        timed_out = set([os.path.basename(fn) for fn, _, _, _ in job_args]) - non_timed_out
        for timed_out_fn in timed_out:
            _add_file_status(
                file_status,
                prefix,
                timed_out_fn,
                "",
                "Timeout",
                "")
            if len(file_status["input_path"]) > file_status_save_period:
                _append_file_status(file_status, file_status_csv)
                file_status = _get_empty_file_status_dict()

    return file_status


def deskew(fn_in, fn_out):
    osd = pytesseract.image_to_osd(fn_in, output_type=pytesseract.Output.DICT)
    angle = osd['orientation']
    image = io.imread(fn_in)
    rotated = rotate(image, angle, resize=True) * 255
    io.imsave(fn_out, rotated.astype(np.uint8))


def extract_text(fn):
    if get_normalized_ext(fn) == ".doc":
        try:
            return textract.process(fn, language="rus")
        except textract.exceptions.ShellError:
            fn_new = os.path.splitext(fn)[0] + ".rtf"
            shutil.copy(fn, fn_new)
            return textract.process(fn_new, language="rus")

    if get_normalized_ext(fn) == ".eml":
        with open(fn) as stream:
            parser = EmailParser(policy=default)
            message = parser.parse(stream)
        text_content = []
        for part in message.walk():
            if part.get_content_type().startswith('text/plain'):
                text_content.append(part.get_content())
        return '\n\n'.join(text_content)

    if get_normalized_ext(fn) == ".xlsx":
        cells_text = []
        wb = load_workbook(filename=fn)
        for sheet_name in wb.sheetnames:
            worksheet = wb[sheet_name]
            cells_text += [str(cell_val) for row_vals in worksheet.values for cell_val in row_vals if cell_val is not None]
        wb.close()
        return '\n'.join(cells_text)

    return textract.process(fn, language="rus")


def textify_fn(fn, _, output_dir):
    if not is_parseable(fn):
        return "Not parsed", None, None
    try:
        if is_image(fn):
            fn_in_base, fn_in_ext = os.path.splitext(fn)
            fn_out = fn_in_base + "_deskewed" + fn_in_ext
            deskew(fn, fn_out)
            fn = fn_out
        text = extract_text(fn)
        if type(text) is bytes:
            text = text.decode(encoding="utf-8")
        text = text.replace("\n", " ")
        with fs_lock:
            output_fn = _non_conflicting_fn(fn, output_dir)
            with open(output_fn, 'w') as f:
                f.write(text)
        return output_fn, None, None
    except Exception as e:
        return "", type(e).__name__, str(e)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--input_dir",
        type=str,
        help="Directory to operate on",
    )
    parser.add_argument(
        "--op",
        choices=["list_extensions", "flatten", "flatten_and_textify"],
        type=str,
        default="flatten_and_textify"
    )
    parser.add_argument(
        "--output_dir",
        type=str,
        default=None,
        help="Directory to output",
    )
    parser.add_argument(
        "--temp_dir",
        type=str,
        default=None,
        help="Directory to use for creating temporary files in during flattening",
    )
    parser.add_argument(
        "--file_status_name",
        type=str,
        default="textify_file_status.csv",
        help="CSV filename to store status of extracted files to",
    )
    parser.add_argument(
        "--file_status_save_period",
        type=int,
        default=1000,
        help="How often to append to the file status CSV file"
    )
    parser.add_argument(
        "--do_not_unzip",
        action="store_true",
        help="Do not unzip any zip files in the input_dir",
    )
    parser.add_argument(
        "--force_restart",
        action="store_true",
        help="Do not skip files already in file status csv",
    )
    parser.add_argument(
        "--num_processes",
        type=int,
        default=10,
        help="Number of parallel processes to use for extraction",
    )
    parser.add_argument(
        "--extraction_timeout",
        type=int,
        default=600,
        help="Number of seconds for extraction timeout",
    )
    parser.add_argument(
        "--verbose",
        action="store_true",
        help="Print warnings during processing in addition to log file",
    )
    args = parser.parse_args()

    if args.verbose:
        logger.setLevel(logging.INFO)
        consoleHandler.setLevel(logging.INFO)
    else:
        warnings.filterwarnings("ignore")

    if not os.path.isdir(args.input_dir):
        raise Exception("Not a directory: " + args.input_dir)

    failed_out_dir = args.output_dir if args.output_dir is not None else '.'
    file_status_csv = os.path.join(failed_out_dir, args.file_status_name)

    if args.op == "list_extensions":
        all_exts = set()
        def _add_ext(fn):
            all_exts.add(get_normalized_ext(fn))
            return "", None, None

        process_dir(
            args.input_dir,
            args.output_dir,
            lambda fn, _1, _2: _add_ext(fn),
            num_pool_procs=args.num_processes,
            pool_timeout=args.extraction_timeout,
            do_not_unzip=args.do_not_unzip)
        print("")
        print("Found extensions:")
        print("=================")
        for ext in sorted(list(all_exts)):
            print(ext)
    elif args.op == "flatten":
        def safe_copy(fn, _, output_dir):
            with fs_lock:
                output_fn = _non_conflicting_fn(fn, output_dir)
                shutil.copy(fn, output_fn)
            return output_fn, None, None

        process_dir(
            args.input_dir,
            args.output_dir,
            lambda fn, _, output_dir: safe_copy(fn, output_dir),
            num_pool_procs=args.num_processes,
            pool_timeout=args.extraction_timeout,
            do_not_unzip=args.do_not_unzip)
    elif args.op == "flatten_and_textify":
        logger.info("Loading already processed files from file status csv...")
        already_processed = set()
        if not args.force_restart and os.path.isfile(file_status_csv):
            already_processed = set(
                pd \
                    .read_csv(file_status_csv) \
                    .fillna("") \
                    .apply(lambda row: row.input_path + row.input_filename, axis=1).tolist()
            )

        file_status = process_dir(
            args.input_dir,
            args.output_dir,
            textify_fn,
            num_pool_procs=args.num_processes,
            pool_timeout=args.extraction_timeout,
            do_not_unzip=args.do_not_unzip,
            already_processed=already_processed,
            file_status_csv=file_status_csv,
            file_status_save_period=args.file_status_save_period)
    else:
        raise ValueError("Unknown operation: " + args.op)

    if args.op == "flatten_and_textify":
        pd \
            .DataFrame(file_status) \
            .to_csv(file_status_csv, index=False)


if __name__ == "__main__":
    main()